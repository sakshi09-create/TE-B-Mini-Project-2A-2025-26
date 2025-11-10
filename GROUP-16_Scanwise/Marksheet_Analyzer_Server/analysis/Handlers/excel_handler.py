import pandas as pd
import re
import openpyxl
from openpyxl.styles import Font
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend for server
import matplotlib.pyplot as plt
import numpy as np
import os


def process_excel_main(input_path, output_path):
    """
    Main function that orchestrates the Excel processing workflow
    """
    print("🔄 Starting Excel processing...")
    
    # Step 1: Process data and calculate percentages
    processed_df = process_data_and_percentages(input_path, output_path)
    
    # Step 2: Highlight names where Remark == 'F'
    highlight_failed_students(output_path)
    
    print(f"✅ Final file saved at: {output_path}")
    return output_path

def process_data_and_percentages(input_path, output_path):
    """
    Read Excel, calculate percentages, and save processed data
    """
    print("🔄 Reading Excel file...")
    df_original = pd.read_excel(input_path, header=None)
    df_data = pd.read_excel(input_path, header=4)

    print(f"Processing {len(df_data)} rows of data...")

    # Identify key columns
    exam_total_col = 'ExamTotal'
    outof_col = 'OUTOF'

    percentages = []
    exam_totals = []

    for idx, value in enumerate(df_data[exam_total_col]):
        if pd.isna(value):
            percentages.append('')
            exam_totals.append('')
            continue

        val_str = str(value).strip()
        parts = val_str.split()

        if len(parts) >= 2:
            first_part = parts[0]
            second_part = parts[-1]
            numeric_score = re.sub(r'@.*', '', second_part)

            if first_part == '--':
                try:
                    outof_value = float(df_data.iloc[idx][outof_col])
                    score_value = float(numeric_score)
                    calculated_perc = (score_value / outof_value) * 100
                    percentages.append(f"{calculated_perc:.2f}%")
                except:
                    percentages.append('--')
            else:
                percentages.append(first_part)

            exam_totals.append(numeric_score)
        else:
            percentages.append('--')
            exam_totals.append('')

    # Insert Percentage column before ExamTotal
    exam_col_pos = df_data.columns.get_loc(exam_total_col)
    new_columns = list(df_data.columns)
    new_columns.insert(exam_col_pos, 'Percentage')

    new_df = pd.DataFrame()
    for i, col in enumerate(new_columns):
        if col == 'Percentage':
            new_df[col] = percentages
        elif col == exam_total_col:
            new_df[col] = exam_totals
        else:
            new_df[col] = df_data[col]

    # Reconstruct full file (metadata + header + data)
    final_rows = []

    for i in range(4):
        row = df_original.iloc[i].tolist()
        while len(row) < len(new_df.columns):
            row.append(None)
        final_rows.append(row)

    header_row = list(new_df.columns)
    final_rows.append(header_row)

    for i in range(len(new_df)):
        final_rows.append(new_df.iloc[i].tolist())

    final_df = pd.DataFrame(final_rows)
    final_df.to_excel(output_path, index=False, header=False)

    print("✅ Data processed and saved.")
    return new_df

def highlight_failed_students(output_path):
    """
    Highlight names in RED where Remark == 'F'
    """
    workbook = openpyxl.load_workbook(output_path)
    sheet = workbook.active

    header_row = None
    name_col = None
    remark_col = None

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        values = [str(cell.value).strip().lower() if cell.value else "" for cell in row]
        if "name" in values and "remark" in values:
            header_row = row[0].row
            for cell in row:
                val = str(cell.value).strip().lower() if cell.value else ""
                if val == "name":
                    name_col = cell.column
                if val == "remark":
                    remark_col = cell.column
            break

    if header_row and name_col and remark_col:
        count = 0
        for row in range(header_row + 1, sheet.max_row + 1):
            remark_value = str(sheet.cell(row=row, column=remark_col).value).strip().upper()
            if remark_value == "F":
                name_cell = sheet.cell(row=row, column=name_col)
                name_cell.font = Font(color="FFFF0000")  # red font
                count += 1

        workbook.save(output_path)
        print(f"🎯 Highlighted {count} names in red where Remark = 'F'.")
    else:
        print("❌ Could not detect 'Name' or 'Remark' headers.")

def process_excel_file(input_path, output_path):
    """
    Main function to process the Excel file.
    Calls helper functions to split, process, and highlight data.
    """
    print("🔄 Starting Excel processing...")
    
    # Step 1: Process data and create new columns
    process_and_split_columns(input_path, output_path)
    
    # Step 2: Highlight names where Remark = 'F'
    highlight_failed_students(output_path)
    
    print(f"✅ Final file saved at: {output_path}")

def process_and_split_columns(input_path, output_path):
    """
    Reads Excel file, processes ExamTotal column to extract percentage and score,
    and saves the modified data to output file.
    """
    print("🔄 Reading Excel file...")
    df_original = pd.read_excel(input_path, header=None)
    df_data = pd.read_excel(input_path, header=4)

    print(f"Processing {len(df_data)} rows of data...")

    # Identify key columns
    exam_total_col = 'ExamTotal'
    outof_col = 'OUTOF'

    percentages = []
    exam_totals = []

    for idx, value in enumerate(df_data[exam_total_col]):
        if pd.isna(value):
            percentages.append('')
            exam_totals.append('')
            continue

        val_str = str(value).strip()
        parts = val_str.split()

        if len(parts) >= 2:
            first_part = parts[0]
            second_part = parts[-1]
            numeric_score = re.sub(r'@.*', '', second_part)

            if first_part == '--':
                try:
                    outof_value = float(df_data.iloc[idx][outof_col])
                    score_value = float(numeric_score)
                    calculated_perc = (score_value / outof_value) * 100
                    percentages.append(f"{calculated_perc:.2f}%")
                except:
                    percentages.append('--')
            else:
                percentages.append(first_part)

            exam_totals.append(numeric_score)
        else:
            percentages.append('--')
            exam_totals.append('')

    # Insert Percentage column before ExamTotal
    exam_col_pos = df_data.columns.get_loc(exam_total_col)
    new_columns = list(df_data.columns)
    new_columns.insert(exam_col_pos, 'Percentage')

    new_df = pd.DataFrame()
    for i, col in enumerate(new_columns):
        if col == 'Percentage':
            new_df[col] = percentages
        elif col == exam_total_col:
            new_df[col] = exam_totals
        else:
            new_df[col] = df_data[col]

    # Reconstruct full file (metadata + header + data)
    final_rows = []

    for i in range(4):
        row = df_original.iloc[i].tolist()
        while len(row) < len(new_df.columns):
            row.append(None)
        final_rows.append(row)

    header_row = list(new_df.columns)
    final_rows.append(header_row)

    for i in range(len(new_df)):
        final_rows.append(new_df.iloc[i].tolist())

    final_df = pd.DataFrame(final_rows)
    final_df.to_excel(output_path, index=False, header=False)

    print("✅ Data processed and saved.")

def highlight_failed_students(output_path):
    """
    Highlights names in RED where Remark == 'F'
    """
    workbook = openpyxl.load_workbook(output_path)
    sheet = workbook.active

    header_row = None
    name_col = None
    remark_col = None

    # Find header row and column indices
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        values = [str(cell.value).strip().lower() if cell.value else "" for cell in row]
        if "name" in values and "remark" in values:
            header_row = row[0].row
            for cell in row:
                val = str(cell.value).strip().lower() if cell.value else ""
                if val == "name":
                    name_col = cell.column
                if val == "remark":
                    remark_col = cell.column
            break

    if header_row and name_col and remark_col:
        count = 0
        for row in range(header_row + 1, sheet.max_row + 1):
            remark_value = str(sheet.cell(row=row, column=remark_col).value).strip().upper()
            if remark_value == "F":
                name_cell = sheet.cell(row=row, column=name_col)
                name_cell.font = Font(color="FFFF0000")  # red font
                count += 1

        workbook.save(output_path)
        print(f"🎯 Highlighted {count} names in red where Remark = 'F'.")
    else:
        print("❌ Could not detect 'Name' or 'Remark' headers.")

def analyze_pass_fail(input_path, chart_output_path):
    """
    Main function for pass/fail analysis.
    Calls helper functions to extract data and generate chart.
    """
    print("🔄 Starting pass/fail analysis...")
    
    # Extract pass/fail data from Excel
    chart_data = extract_pass_fail_data(input_path)
    
    # Generate chart image
    generate_pass_fail_chart(chart_data, chart_output_path)
    
    print(f"✅ Chart saved at: {chart_output_path}")
    
    return chart_data

def extract_pass_fail_data(input_path):
    """
    Extracts pass/fail counts for each subject from the Excel file.
    """
    # Load Excel (row 6 and 7 are headers)
    df = pd.read_excel(input_path, sheet_name="Sheet1", header=[5, 6])
    
    # Flatten the multi-index column names
    df.columns = [
        '_'.join([str(c) for c in col if str(c) != 'nan']).strip()
        for col in df.columns
    ]
    
    # Print columns for debugging
    print("Available columns (first 40):")
    print(df.columns.tolist()[:40])
    
    # Manually specify the grade columns you want to analyze
    subjects = {
        "COURSE-1_SE": "COURSE-1_SE.1",
        "COURSE-2_SE": "COURSE-2_SE.1",
        "COURSE-3_SE": "COURSE-3_SE.1",
        "COURSE-4_SE": "COURSE-4_SE.1",
        "COURSE-5_SE": "COURSE-5_SE.1",
        "COURSE-6_SE": "COURSE-6_SE.1",
    }
    
    # Prepare data
    pass_counts = []
    fail_counts = []
    courses = []
    
    for course, grade_col in subjects.items():
        if grade_col not in df.columns:
            print(f"⚠️ Skipping {course} — column '{grade_col}' not found.")
            continue
        
        results = df[grade_col].dropna().astype(str).str.strip().str.upper()
        
        total = len(results)
        if total == 0:
            continue
        
        failed = (results == "F").sum()
        passed = total - failed
        
        pass_counts.append(int(passed))
        fail_counts.append(int(failed))
        courses.append(course)
    
    chart_data = {
        "courses": courses,
        "pass_counts": pass_counts,
        "fail_counts": fail_counts
    }
    
    print(f"✅ Extracted data for {len(courses)} courses")
    
    return chart_data

def generate_pass_fail_chart(chart_data, output_path):
    """
    Generates a stacked bar chart showing pass/fail counts per course.
    """
    courses = chart_data["courses"]
    pass_counts = chart_data["pass_counts"]
    fail_counts = chart_data["fail_counts"]
    
    # Create the plot
    plt.figure(figsize=(10, 6))
    plt.bar(courses, pass_counts, label="Pass Count", color="#7dc87dff")
    plt.bar(courses, fail_counts, bottom=pass_counts, label="Fail Count", color="#C97C7C")
    
    # Add count labels on each bar
    for i, course in enumerate(courses):
        plt.text(i, pass_counts[i] / 2, f"{pass_counts[i]}", ha="center", color="white", weight="bold")
        plt.text(i, pass_counts[i] + fail_counts[i] / 2, f"{fail_counts[i]}", ha="center", color="white", weight="bold")
    
    plt.xlabel("Courses")
    plt.ylabel("Number of Students")
    plt.title("Pass vs Fail Count per Course")
    plt.legend()
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()
    
    # Save to file
    plt.savefig(output_path, dpi=150, bbox_inches='tight')
    plt.close()
    
    print(f"✅ Chart generated successfully")

def calculate_semester_average(input_paths, output_path):
    """
    Main function to calculate average percentage across multiple semesters.
    Calls helper functions to process each file and merge results.
    """
    print(f"🔄 Starting semester average calculation for {len(input_paths)} semesters...")
    
    # Process each semester file
    dfs = []
    for i, fpath in enumerate(input_paths):
        print(f"Processing semester {i+1}: {os.path.basename(fpath)}")
        df = preprocess_semester_df(fpath, i+1)
        dfs.append(df)
    
    # Merge all dataframes
    merged_df = merge_semester_dfs(dfs, len(input_paths))
    
    # Save to output
    merged_df.to_excel(output_path, index=False)
    
    print(f"✅ Semester average file saved at: {output_path}")

def clean_percentage_value(val):
    """Convert percentage string like '61.93%' or '--' to float or NaN."""
    try:
        if pd.isna(val):
            return np.nan
        txt = str(val).strip()
        if txt == '--' or txt == '':
            return np.nan
        txt = txt.replace('%', '')
        return float(txt)
    except Exception:
        return np.nan

def find_column_by_keywords(df, keywords):
    """Find column in dataframe that matches any of the given keywords."""
    for col in df.columns:
        col_norm = col.replace(" ", "").upper()
        for kw in keywords:
            if kw in col_norm:
                return col
    return None

def preprocess_semester_df(fpath, perc_index):
    """
    Reads and preprocesses a single semester Excel file.
    Extracts Roll No, Name, and Percentage columns.
    """
    # Always read with header at row 5 (0-indexed row 4)
    df = pd.read_excel(fpath, header=4)
    df.columns = df.columns.str.strip()
    
    print(f"\nColumns in file {os.path.basename(fpath)}: {list(df.columns)}")
    
    # Flexible detection for Roll No/Name/Percentage
    roll_col = find_column_by_keywords(df, ["ROLL"])
    name_col = find_column_by_keywords(df, ["NAME"])
    perc_col = find_column_by_keywords(df, ["PERCENTAGE"])
    
    if roll_col is None or name_col is None or perc_col is None:
        raise ValueError(
            f"File {os.path.basename(fpath)} missing Roll No or Name or Percentage columns.\n"
            f"Columns found: {list(df.columns)}"
        )
    
    # Rename columns to standard names
    df = df.rename(columns={
        roll_col: 'ROLLNO',
        name_col: 'NAME',
        perc_col: f'Percentage {perc_index}'
    })
    
    # Normalize Roll No and Name keys for joining
    df['ROLLNO_PRE'] = df['ROLLNO'].astype(str).str.upper().str.strip()
    df['NAME_PRE'] = df['NAME'].astype(str).str.upper().str.strip()
    
    # Clean percentage
    perc_col_name = f'Percentage {perc_index}'
    df[perc_col_name] = df[perc_col_name].apply(clean_percentage_value)
    
    # Keep only necessary columns to avoid merge conflicts
    keep_cols = ['ROLLNO', 'NAME', 'ROLLNO_PRE', 'NAME_PRE', perc_col_name]
    df = df[keep_cols]
    
    print(f"Loaded {os.path.basename(fpath)} with {len(df)} records.")
    
    return df

def merge_semester_dfs(dfs, n_sem):
    """
    Merges multiple semester dataframes on Roll No and Name.
    Calculates average percentage across all semesters.
    """
    # Start with first dataframe
    merged = dfs[0]
    
    # Merge remaining dataframes one by one
    for i in range(1, len(dfs)):
        merged = pd.merge(
            merged, dfs[i],
            on=['ROLLNO_PRE', 'NAME_PRE'],
            how='outer',
            suffixes=('', '_right')
        )
        
        # Fix ROLLNO and NAME if suffixed
        for col in ['ROLLNO', 'NAME']:
            right_col = col + '_right'
            if right_col in merged.columns:
                if col in merged.columns:
                    merged[col] = merged[col].combine_first(merged[right_col])
                else:
                    merged[col] = merged[right_col]
                merged = merged.drop(right_col, axis=1)
    
    # Drop the _PRE columns after merge
    merged = merged.drop(['ROLLNO_PRE', 'NAME_PRE'], axis=1)
    
    # Calculate average percentage
    percentage_cols = [f'Percentage {i+1}' for i in range(n_sem)]
    merged['Average Percentage'] = merged[percentage_cols].mean(axis=1, skipna=True).round(2)
    
    # Prepare final output columns
    output_columns = ['ROLLNO', 'NAME'] + percentage_cols + ['Average Percentage']
    final_df = merged[output_columns]
    
    print(f"✅ Merged {n_sem} semesters with {len(final_df)} total students")
    
    return final_df
